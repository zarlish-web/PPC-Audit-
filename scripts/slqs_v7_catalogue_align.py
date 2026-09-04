import pandas as pd, numpy as np, shutil, openpyxl, json, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
SRC='SLQS_Decided_Bulk_v6_CORRECTED_04Sep2026.xlsx'
OUT='SLQS_Decided_Bulk_v7_SKU_ALIGNED_04Sep2026.xlsx'
shutil.copy(SRC,OUT)
num=lambda s: pd.to_numeric(s,errors='coerce')
RX=r'twin xl|cal king|california king|super king|full size|full bedspread|bedspreads full|bed quilt full|olive green|light blue|winter|with sheets'
def why(t):
    t=str(t).lower()
    if re.search(r'twin xl',t): return 'SIZE','No Twin XL SKU. We sell Twin, Queen, King only.'
    if re.search(r'cal king|california king',t): return 'SIZE','No California King SKU.'
    if re.search(r'super king',t): return 'SIZE','No Super King SKU.'
    if re.search(r'full size|full bedspread|bedspreads full|bed quilt full',t): return 'SIZE','No Full SKU.'
    if 'olive green' in t: return 'COLOUR','Our only green is Sage Green.'
    if 'light blue' in t: return 'COLOUR','Our only blue is Navy Blue.'
    if 'winter' in t: return 'SEASON','Product is a lightweight summer bedspread.'
    if 'with sheets' in t: return 'BUNDLE','Set is quilt plus shams. No sheets included.'
    return None,None

wb=openpyxl.load_workbook(OUT)
neg=[]

# ---------- 1. strip bad keyword rows from the proposed builds ----------
for sh in ['New SP Campaigns','New SB Campaigns']:
    ws=wb[sh]; hdr=[c.value for c in ws[1]]
    ci_kw=hdr.index('KeywordText')+1; ci_cn=hdr.index('CampaignName')+1
    camp=None; drop=[]
    for r in range(2,ws.max_row+1):
        cn=ws.cell(row=r,column=ci_cn).value
        if cn: camp=cn
        t=ws.cell(row=r,column=ci_kw).value
        if t and re.search(RX,str(t).lower()):
            cat,w=why(t); drop.append(r); neg.append((str(t),cat,w,camp or '',sh,'Removed before launch'))
    for r in reversed(drop): ws.delete_rows(r)
    print(f"{sh:20s} removed {len(drop)} keyword rows")

# ---------- 2. negate + pause the live ones ----------
fb=pd.read_excel(SRC,sheet_name='Final Bulk')
P='Portfolio Name (Informational only)'; CN='Campaign Name (Informational only)'
mask=(fb[P]=='Quilt Set')&(fb['Entity']=='Keyword')&fb['Keyword Text'].astype(str).str.lower().str.contains(RX,na=False)
ws=wb['Final Bulk']; hdr=[c.value for c in ws[1]]
c_act=hdr.index('Action')+1; c_rsn=hdr.index('Reasoning')+1; c_nb=hdr.index('New Bids')+1
amber=PatternFill('solid',fgColor='FFF2CC')
for i in fb[mask].index:
    t=fb.at[i,'Keyword Text']; cat,w=why(t); r=i+2
    ws.cell(row=r,column=c_act,value='Pause').fill=amber
    ws.cell(row=r,column=c_rsn,value=(
        f'{cat} MISMATCH — not a performance decision. "{t}" describes a product we do not sell. {w} '
        'Negated on sight at any click count, the same way an irrelevant term is: no bid makes a click convert '
        'when the shopper has specified something our catalogue does not contain. Added as negative exact in this '
        'campaign so the search term cannot re-enter through another keyword. Reverses only if the SKU range changes.'
    )).fill=amber
    ws.cell(row=r,column=c_nb,value=None).fill=amber
    neg.append((str(t),cat,w,str(fb.at[i,CN]),'Final Bulk','Paused and negated'))
print(f"{'Final Bulk':20s} paused+negated {mask.sum()} live keywords")

# ---------- 3. the Negatives tab (was a named gap) ----------
navy=PatternFill('solid',fgColor='1F3864'); thin=Side(style='thin',color='B7C0D0')
ng=wb.create_sheet('Negatives',3)
rows=[('Negative keyword','Match type','Class','Why it does not fit','Campaign','Status')]
for t,cat,w,camp,src,st in sorted(neg):
    rows.append((t,'negativeExact',cat,w,camp,st))
for i,row in enumerate(rows,1):
    for j,v in enumerate(row,1):
        c=ng.cell(row=i,column=j,value=v)
        c.font=Font(name='Arial',size=9.5,bold=(i==1),color='FFFFFF' if i==1 else '000000')
        c.alignment=Alignment(wrap_text=True,vertical='top'); c.border=Border(left=thin,right=thin,top=thin,bottom=thin)
        if i==1: c.fill=navy
for col,wd in zip('ABCDEF',[36,14,10,52,28,22]): ng.column_dimensions[col].width=wd
ng.freeze_panes='A2'

# ---------- 4. the SKU matrix, so the check is reproducible ----------
skus=json.load(open('skus.json'))
sm=wb.create_sheet('SKU Matrix',4)
sm.append(['SIZES WE SELL','TWIN','QUEEN','KING','','COLOURS WE SELL','BLACK','IVORY','LIGHT GREY','NAVY BLUE','SAGE GREEN','TAUPE','WHITE'])
sm.append([])
sm.append(['Size','Colour','ASIN','Available','Status','Pieces'])
for s,c,a,q,st in skus:
    sm.append([s,c.replace('-',' '),a,q,st,'2 Pc (1 sham)' if s=='TWIN' else '3 Pc (2 shams)'])
sm.append([])
sm.append(['NOT IN THE CATALOGUE — negate on sight','Twin XL, Full, Full/Queen, California King, Super King, Split King, Crib, Toddler, Daybed'])
sm.append(['','Pink, purple, red, yellow, orange, brown, cream, beige, tan, gold, silver, dark grey, charcoal, olive/emerald/forest green, light/sky/baby/royal blue, teal, turquoise'])
sm.append(['','Sheets, comforter-only intent, duvet, blanket, mattress pad, bed skirt, weighted, heated, winter/thermal/fleece'])
for i in range(1,sm.max_row+1):
    for c in sm[i]:
        c.font=Font(name='Arial',size=9.5,bold=(i in (1,3) or (c.column==1 and i>sm.max_row-3)))
        c.alignment=Alignment(wrap_text=True,vertical='top')
        if i in (1,3): c.fill=navy; c.font=Font(name='Arial',size=9.5,bold=True,color='FFFFFF')
for col,wd in zip('ABCDEF',[26,18,16,12,16,16]): sm.column_dimensions[col].width=wd
wb.save(OUT)
print(f"\nsaved {OUT}   negatives listed: {len(neg)}")
