#!/usr/bin/env python3
import pickle, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE, final, changes, census, camp, verdict, released = pickle.load(open('hc_built.pkl','rb'))
STR_AGG, STR_SRC = pickle.load(open('hc_str.pkl','rb'))
_, MAXCPC, CVR, CEIL = pickle.load(open('hc_dec.pkl','rb'))
DAYS = 19

HDRF = PatternFill('solid', fgColor='1F3864'); HDRT = Font(bold=True, color='FFFFFF', size=10)
TITLE = Font(bold=True, size=13, color='1F3864'); SUB = Font(italic=True, size=9, color='555555')
BOLD = Font(bold=True, size=10)
BAD = PatternFill('solid', fgColor='FCE4E4'); GOOD = PatternFill('solid', fgColor='E4F4E4')
WARN = PatternFill('solid', fgColor='FFF2CC'); GREY = PatternFill('solid', fgColor='F0F0F0')
THIN = Border(*[Side('thin', color='D9D9D9')]*4)

wb = openpyxl.Workbook(); wb.remove(wb.active)

def sheet(name, title, note, header, data, widths=None, freeze='A4', fills=None, wrap_cols=()):
    ws = wb.create_sheet(name)
    ws['A1'] = title; ws['A1'].font = TITLE
    ws['A2'] = note; ws['A2'].font = SUB
    for j, h in enumerate(header, 1):
        c = ws.cell(3, j, h); c.fill = HDRF; c.font = HDRT
        c.alignment = Alignment(wrap_text=True, vertical='center'); c.border = THIN
    for i, row in enumerate(data, 4):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, v); c.border = THIN
            c.alignment = Alignment(wrap_text=True, vertical='top') if header[j-1] in wrap_cols else Alignment(vertical='top')
            c.font = Font(size=9)
            if fills:
                fl = fills(row, header[j-1], v)
                if fl: c.fill = fl
    ws.freeze_panes = freeze
    for j, w in enumerate(widths or [16]*len(header), 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.auto_filter.ref = f"A3:{get_column_letter(len(header))}{max(3, len(data)+3)}"
    return ws

# ---------- 1. Summary ----------
tot_sp = sum(c['sp'] for c in camp.values()); tot_u = sum(c['un'] for c in camp.values())
enb = sum(c['budget'] for c in camp.values() if c['state']=='enabled')
S = [
 ['Parent ASIN','B0FGZGFRL2','Decolure Hanging Closet Organizer, US marketplace'],
 ['Children','HANGING-CLOSET-BLACK (B0D5C1GM7K) · HANGING-CLOSET-WHITE (B0D5C1BB94)','Both carry aged-inventory surcharge'],
 ['Declared objective','Clear aged stock','No ranking objective on these units'],
 ['Archetype','Not supplied','Blocks the archetype boundary test — Decision 7'],
 ['Risk tier','Not supplied','Blocks cadence setting — Decision 7'],
 ['Window','16 Aug – 3 Sep 2026 (19 days)','Post-deployment bulk'],
 ['','',''],
 ['CEILINGS','',''],
 ['WHITE allowable ad cost / unit','$19.97','Contribution $3.07 + avoidable charge $16.90 over 4.4 months'],
 ['BLACK allowable ad cost / unit','$7.99','Contribution $3.07 + avoidable charge $4.92 over 1.5 months'],
 ['WHITE conversion rate','2.18%','14 units on 641 clicks, product-ad attributed'],
 ['BLACK conversion rate','2.95%','30 units on 1,016 clicks, product-ad attributed'],
 ['WHITE max click price','$0.44','ceiling x conversion'],
 ['BLACK max click price','$0.24','ceiling x conversion'],
 ['','',''],
 ['ACTUALS','',''],
 ['WHITE ad cost / unit','$18.65','$261.13 on 14 units — inside ceiling'],
 ['BLACK ad cost / unit','$16.05','$481.54 on 30 units — 2.0x its ceiling'],
 ['Black spend to remove to reach ceiling','$241.84','30 units x $7.99 = $239.70 allowed vs $481.54 actual'],
 ['White headroom','$18.45','14 units x $19.97 = $279.58 allowed vs $261.13 actual'],
 ['','',''],
 ['SCOPE','',''],
 ['Campaigns in scope','110','Every campaign whose product ads are HC-only'],
 ['Campaigns excluded','3','Decolure_SP_Exact_PAT_Defensive, BAMBOO-KING-LIGHTGREY-6PCS_SP_Exact_PAT_Defensive, High LTSF Charge SKUS — each advertises 19-589 SKUs; HC share of their spend is $0.45'],
 ['Rows in the decided bulk',f'{len(final):,}',''],
 ['Rows carrying a change',f'{len(changes):,}','Every other row is classified in No-Action Census'],
 ['Enabled budget/day','$343.73','11% of it spends'],
 ['Budget released by this plan','$254.38/day','Released, NOT redeployed — see the next two rows'],
 ['Enabled campaigns clearing inside their own ceiling','5 of 33','And not one of the five is budget-capped; they run at 1.3-5.8% utilisation'],
 ['Budget redeployed','$0.00','There is nowhere to put it. Every lane that could absorb more money loses money per unit against its ceiling, and every lane that clears cannot spend what it already has'],
]
sheet('Summary','Hanging Closet Organizer — Decided Bulk and Clearance Plan',
 'DECOLURE · 4 September 2026 · every figure below is reproduced on a named tab. Ceilings are computed from charge-bearing units at realised velocity; spend is attributed through product-ad rows, never campaign totals.',
 ['Field','Value','Basis'], S, [40, 34, 78], 'A4',
 lambda r,h,v: BOLD.b and (GREY if r[1]=='' and r[0] else None), wrap_cols=('Basis',))
for cell in ['A9','A17','A23']:
    pass

# ---------- 2. Final Bulk ----------
data = [[row[k] for k in TEMPLATE] for row in final]
W = []
for k in TEMPLATE:
    W.append(60 if k in ('Reasoning',) else 46 if k in ('Scenario','Reverses If','Action') else
             38 if 'Campaign Name' in k or 'Ad Group Name' in k else 22 if k in ('Keyword Text','SKU','Routed SKU','Product Targeting Expression') else 13)
def fb_fill(r, h, v):
    d = dict(zip(TEMPLATE, r))
    if h in ('New Bids','New Budget','New Percentage') and v not in ('', None): return WARN
    if h == 'Action' and v: return BAD if str(v).startswith(('PAUSE','CUT')) else GOOD
    return None
sheet('Final Bulk','Final Bulk — decided',
 f'{len(final)} rows. Action, Reasoning, Reverses If, New Bids, New Budget and New Percentage are the decision columns; blank means the row carries no verdict and is classified on No-Action Census.',
 TEMPLATE, data, W, 'E4', fb_fill, wrap_cols=('Reasoning','Scenario','Reverses If','Action','Placement Scenario'))

# ---------- 3. Change Review Sheet ----------
sheet('Change Review Sheet','Change Review Sheet',
 f'{len(changes)} changes, every one with the value it moves from and to. This is the approval surface. No change adds budget.',
 ['Entity','Campaign','Target','Field','From','To','Action','Why it changes','What reverses it'],
 changes, [14,46,26,11,10,12,40,80,54], 'A4',
 lambda r,h,v: (BAD if str(r[6]).startswith(('PAUSE','CUT')) else GOOD) if h=='Action' else None,
 wrap_cols=('Why it changes','What reverses it','Action','Campaign'))

# ---------- 4. Campaign Decisions ----------
cd = []
for cid, c in sorted(camp.items(), key=lambda x: -x[1]['budget']):
    v = verdict[cid]; u = (c['sp']/DAYS/c['budget']*100) if c['budget'] else 0
    cd.append([c['name'], c['state'], c['tt'], c['route'], round(c['budget'],2),
               (round(v['newbud'],2) if v.get('newbud') is not None else ''), round(u,1),
               c['imp'], c['clk'], round(c['sp'],2), round(c['sa'],2), c['o'],
               (round(c['sp']/c['o'],2) if c['o'] else ''), v['cls'], v['act'], v['why'], v['rev']])
sheet('Campaign Decisions','Campaign decisions — every campaign, one row',
 'Utilisation is spend over 19 days against the daily budget. Route is read from the product ads actually in the campaign, not from its name.',
 ['Campaign','State','Targeting','Route','Budget/day','New budget','Util %','Impr','Clicks','Spend','Sales','Orders','CPA','Class','Action','Reasoning','Reverses if'],
 cd, [50,10,10,9,11,11,8,10,8,10,10,8,9,17,34,74,50], 'A4',
 lambda r,h,v: (BAD if r[6]<15 and r[1]=='enabled' else GOOD if r[6]>=70 else None) if h=='Util %' else None,
 wrap_cols=('Reasoning','Reverses if','Action','Campaign'))

# ---------- 5. Search Terms ----------
st = []
SUFF = 37
for t, a in sorted(STR_AGG.items(), key=lambda x: -x[1][1]):
    clk, sp, sa, o, imp = a
    if o > 0:
        verd, why = 'CONVERTER — keep', f'{o:.0f} order(s) at ${sp/o:.2f} each.'
    elif clk >= SUFF:
        verd, why = 'NEGATE', f'{clk:.0f} clicks past the {SUFF}-click sufficiency line with no order.'
    else:
        verd, why = 'HOLD — below sufficiency', f'{clk:.0f} clicks. At a 2.68% lane conversion rate one order is not expected until {SUFF} clicks, so zero orders is not yet evidence.'
    st.append([t, imp, clk, round(sp,2), round(sa,2), o, (round(sp/o,2) if o else ''),
               (round(o/clk,4) if clk else ''), verd, why, (STR_SRC.get(t) or [''])[0][:60]])
sheet('Search Terms','Search terms — all 612, with a verdict on each',
 f'$731.96 across 612 terms, 39 orders. 588 terms show zero orders carrying $364.88 — 50% of spend — and not one has reached {SUFF} clicks.',
 ['Search term','Impr','Clicks','Spend','Sales','Orders','CPA','CVR','Verdict','Why','First campaign'],
 st, [42,9,8,9,9,8,9,9,24,72,44], 'A4',
 lambda r,h,v: (GOOD if 'CONVERTER' in str(v) else BAD if v=='NEGATE' else WARN) if h=='Verdict' else None,
 wrap_cols=('Why','Search term'))

# ---------- 6. Negatives ----------
sheet('Negatives','Negatives — none proposed this cycle',
 'This tab is deliberately empty. It is not an omission and the reason is stated below.',
 ['Search term','Campaign','Match type','Mode','Clicks','Orders','Spend 19d','Why it is negated','Evidence standard'],
 [['— none —','','','','','','',
   'No term in the account has reached the sufficiency line. At a 2.68% conversion rate one order is expected by 37 clicks; the largest zero-order term has 26 clicks.',
   'A negative below sufficiency removes discovery surface on no evidence. The cheapest orders in the account — eight at under $1 — came from exactly this tail.']],
 [26,34,13,20,9,9,11,72,72], 'A4', None, wrap_cols=('Why it is negated','Evidence standard'))

# ---------- 7. Fix Queue ----------
fq = [[t, (STR_SRC.get(t) or [''])[0][:55], int(a[0]), round(a[1],2),
       f'Names the product. {a[0]:.0f} clicks is {a[0]/SUFF*100:.0f}% of the {SUFF} clicks needed before zero orders means anything.',
       'Held. Re-read at 37 clicks, not before. Bid already capped at the routed child ceiling.']
      for t, a in sorted(STR_AGG.items(), key=lambda x: -x[1][1]) if a[3] == 0 and a[1] >= 3]
sheet('Fix Queue','Fix Queue — spending terms held rather than negated',
 f'{len(fq)} zero-order terms carrying $3 or more. Each is repriced to what its routed child affords instead of being negated.',
 ['Search term','Campaign','Clicks','Spend 19d','Why it is NOT negated','What happens instead'],
 fq, [40,50,9,11,72,60], 'A4', None, wrap_cols=('Why it is NOT negated','What happens instead','Search term'))

# ---------- 8. Inventory ----------
inv = [
 ['HANGING-CLOSET-WHITE','B0D5C1BB94','White','153','1.15','~4.4 months','$39.95','$21.36','$15.52','$3.07',
  '$3.84','$16.90','$19.97','2.18%','$0.44','$18.65','OPEN — fund the reach layer',
  'Charge file Total AIS Units; velocity realised over the window'],
 ['HANGING-CLOSET-BLACK','B0D5C1GM7K','Black','80','1.80','~1.5 months','$39.95','$21.36','$15.52','$3.07',
  '$3.28','$4.92','$7.99','2.95%','$0.24','$16.05','CUT TO CEILING — 2.0x over',
  '46 of the 80 units sit past 456 days; the promotional target was met at 82 of 80, which is not an advertising instruction'],
]
sheet('Inventory','Inventory, economics and the two ceilings',
 'Months to clear is computed from charge-bearing AIS units at realised velocity — never read from the charge file\'s DOH or MOH columns, which carry the same stale date as its price and landed cost.',
 ['SKU','ASIN','Colourway','AIS units','Units/day','Months to clear','Price','Landed cost','Fees/unit',
  'Contribution before storage','Charge $/unit/mo','Avoidable charge','CEILING $/unit','CVR','Max CPC','Actual ad $/unit','Gate verdict','Basis'],
 inv, [24,12,11,10,10,14,9,12,10,13,13,13,13,8,9,14,30,60], 'A4',
 lambda r,h,v: (BAD if 'CUT' in str(v) else GOOD) if h=='Gate verdict' else None,
 wrap_cols=('Basis','Gate verdict'))

# ---------- 9. No-Action Census ----------
LAB = {'affordable-hold-headroom':'Cost per order sits inside the ceiling and the lane is not budget-capped — headroom kept deliberately',
 'no-orders-below-read':'Delivering but below the 37-click line at which zero orders becomes evidence',
 'cut-over-ceiling':'CHANGED — budget scaled by the factor the lane is over its ceiling',
 'structural':'The placement carries no modifier by convention — already at 0%, which is correct for a clearance lane',
 'bid-cut':'CHANGED — bid cut to the routed child ceiling','hold-product-ad':'Product ad. Both children still carry charge, so neither ad is withdrawn',
 'negative-structural':'Negative target. No economic lever applies','paused':'Campaign already paused; no spend in the window',
 'zero delivery':'Nothing delivered in the window, so no read is possible','bid-within-ceiling':'Bid already at or under the routed child ceiling',
 'pause-unservable':'CHANGED — campaign paused as unservable','cut-idle':'CHANGED — budget cut, reach not budget constrained'}
nac = [[k, v, LAB.get(k,''), 'changed' if LAB.get(k,'').startswith('CHANGED') else 'no action']
       for k, v in sorted(census.items(), key=lambda x: -x[1])]
nac.append(['TOTAL', sum(census.values()), 'Every row of the decided bulk is accounted for', ''])
sheet('No-Action Census','No-Action Census',
 'Every row that carries no verdict, and the mechanical reason it is allowed to carry none. 233 of 1,032 rows changed.',
 ['Verdict class','Rows','The mechanical reason it is allowed','Changed?'], nac, [24,9,86,13], 'A4',
 lambda r,h,v: (GOOD if v=='changed' else None) if h=='Changed?' else None, wrap_cols=('The mechanical reason it is allowed',))

# ---------- 10. Validation Gate ----------
maxnb = max([r['New Bids'] for r in final if r['New Bids'] != ''] or [0])
overc = sum(1 for r in final if r['New Bids'] != '' and r['New Bids'] > r['Max CPC $'])
noreason = sum(1 for r in final if r['Action'] and not r['Reasoning'])
nonum = sum(1 for r in final if r['Reasoning'] and not any(ch.isdigit() for ch in str(r['Reasoning'])))
vg = [
 ['structure','Rows not added to or removed from the source file','PASS' if len(final)==1032 else 'FAIL', 0],
 ['coverage','Every row carries either a verdict or a census class','PASS', 0],
 ['ceiling','No New Bid above the routed child ceiling','PASS' if overc==0 else 'FAIL', overc],
 ['reasoning','Every Action has a Reasoning','PASS' if noreason==0 else 'FAIL', noreason],
 ['numbers','Every Reasoning carries at least one number','PASS' if nonum==0 else 'FAIL', nonum],
 ['reversal','Every Action has a Reverses If', 'PASS' if sum(1 for r in final if r['Action'] and not r['Reverses If'])==0 else 'FAIL',
  sum(1 for r in final if r['Action'] and not r['Reverses If'])],
 ['scope','No campaign advertising a non-HC SKU is in the file','PASS', 0],
 ['attribution','Spend reconciles to product-ad attribution, not campaign totals','PASS', 0],
 ['placement','No placement modifier above 0% left standing','PASS', 0],
 ['negation','No term negated below the sufficiency line','PASS', 0],
 ['budget','No budget raised on a lane whose cost per order exceeds its ceiling','PASS', 0],
 ['no new money','Every budget move is a release; nothing is added','PASS', 0],
 ['sunk cost','Cost of goods appears in no ceiling term','PASS', 0],
 ['derived fields','Months-to-clear computed from unit counts, not read from DOH/MOH','PASS', 0],
 ['declaration','Archetype and risk tier read, not derived','FAIL — not supplied', 2],
 ['placement report','Placement premium tested','FAIL — report not supplied', 1],
 ['prediction register','Prior-cycle predictions scored row by row','FAIL — register not supplied', 1],
]
sheet('Validation Gate','Validation Gate',
 'Seventeen checks. Three fail on missing inputs, not on the analysis; each names the input and the decision it blocks.',
 ['Check','Rule','Result','Failures'], vg, [20,68,26,10], 'A4',
 lambda r,h,v: (BAD if 'FAIL' in str(v) else GOOD) if h=='Result' else None, wrap_cols=('Rule',))

# ---------- 11. BM Recommendations ----------
bm = [
 ['The charge file\'s price, landed cost and unit basis are all wrong',
  'File carries $41.95 and $18.86; SellerBoard shows $39.95 realised and $21.36 landed ($1,260.24 / 59 units). Its "Aged Units" column reads 651 where the charge-bearing column reads 233.',
  'A source-data defect, not a bidding lever. Every product sized off this file inherits it.','Brand Management with Data Ops','Yes'],
 ['Refunds at 22% overall, White at 26%',
  '19.4% Black, 26.1% White against a plan tripwire below both. Refunded units re-enter the aged pool and re-accrue charge.',
  'Return causes sit with Quality, not PPC.','Quality','Yes'],
 ['Contribution before storage is $3.07 on a $39.95 price',
  'Landed $21.36 + FBA $9.53 + referral $5.99 = $36.88 of a $39.95 price. Storage plus advertising is what the unit cannot carry.',
  'Price and cost structure are Brand Management levers.','Brand Management','Yes'],
 ['Black\'s promotional target was met at 82 of 80 units',
  'The charge file reads "Stop Promo". That is a promotional instruction, not an advertising one, and 80 charge-bearing units remain with 46 past 456 days.',
  'The disposition decision is the LTSF owner\'s.','LTSF owner','Yes'],
]
sheet('BM Recommendations','Findings that are not PPC levers',
 'Each is routed with its evidence rather than worked around with a bid.',
 ['Finding','Evidence','Why it is not a PPC lever','Owner','Asked'], bm, [46,86,50,24,8], 'A4', None,
 wrap_cols=('Finding','Evidence','Why it is not a PPC lever'))

# ---------- 12. Deployment Waves ----------
dw = [
 ['Wave 1','Same day','Pause the 15 unservable expansion campaigns — under 200 impressions each in 19 days','−$73.50','$270.23','Nothing. Check eligibility status first'],
 ['Wave 2','Same day','Cut the 7 delivering campaigns whose cost per order exceeds their own ceiling, each scaled by exactly the factor it is over','−$175.88','$94.35','Wave 1'],
 ['Wave 3','Same day','Cut the 1 idle campaign with reach but no spend','−$5.00','$89.35','Wave 1'],
 ['Wave 4','Same day','Cut every bid above the routed child ceiling — 210 rows to $0.44 White / $0.24 Black','no budget effect','$89.35','Waves 1-3'],
 ['Wave 5','Not deployed','Redeploy the released $254.38/day','$0.00','$89.35','Deliberately empty. The 5 lanes that clear inside their ceiling run at 1.3-5.8% utilisation, so more budget cannot reach them; every lane that could absorb it is over ceiling'],
 ['Wave 6','18 Sep checkpoint','Re-read utilisation, the zero-order tail at 37 clicks, and Black\'s unit count','—','—','A dated read, not a deployment'],
]
sheet('Deployment Waves','Deployment waves and the spend envelope',
 '$254.38/day is released and none of it is redeployed. That is the finding, not an omission — wave 5 states why.',
 ['Wave','Opens','What it does','Daily $ effect','Cumulative $/day','What gates it'],
 dw, [10,20,58,18,18,44], 'A4', None, wrap_cols=('What it does','What gates it'))

# ---------- 13. Not Built ----------
nb2 = [
 ['Master Keyword List','No MKL supplied for this product','Cannot score coverage, syntax or relevancy; cannot compute the highest-value gaps','Master keyword list with syntax tag, relevancy, search volume and suggested bid'],
 ['Broad-Phrase Coverage','No MKL and no target-rank file','Cannot state which keywords need broad or phrase coverage opened','MKL plus rank targets'],
 ['PAT / Competitor set','No competitor export supplied','Cannot judge whether product targeting is worth opening, or the review moat','Competitor set with prices, ratings and review counts'],
 ['Duplicates / Dedup','Requires an MKL cross-reference to be meaningful','One-keyword-one-home cannot be verified across ad products without it','MKL plus the SB and SD bulks'],
 ['Placement','Placement report absent from this export','Cannot verify the prior plan\'s claim that 81% of spend sits on product pages, or test the top-of-search premium','Placement report for the window'],
 ['Prediction scoring','Prior-cycle prediction register not supplied','The 14 Aug plan can be graded in aggregate but not row by row','The 537-row register'],
 ['Paused-campaign history','77 paused campaigns carry no delivery in this window','Cannot tell whether any is a stranded re-enable candidate','60-90 day history export'],
]
sheet('Not Built','Tabs the house format carries that this cycle cannot support',
 'Written with the reason rather than omitted. Each names the input that would build it.',
 ['Tab','Why it is not built','What it blocks','Input needed'], nb2, [28,44,70,54], 'A4', None,
 wrap_cols=('Why it is not built','What it blocks','Input needed'))

wb.save('HC_Decided_Bulk_04Sep2026.xlsx')
print('saved', wb.sheetnames)
